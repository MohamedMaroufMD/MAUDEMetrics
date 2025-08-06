"""
MAUDEMetrics - FDA Medical Device Adverse Event Explorer
Copyright (c) 2025 Mohamed Marouf, MD

This file is part of MAUDEMetrics, an open-source tool for exploring adverse event 
reports submitted to the FDA's MAUDE database, enabling efficient safety signal 
detection and reporting.

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at:
    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.

For research and educational purposes only. Not for clinical decision-making.
"""

from flask import Flask, request, render_template, redirect, url_for, send_file, session, send_from_directory
import requests
import pandas as pd
import sqlite3
from datetime import datetime
import json
import os
import re

app = Flask(__name__)
app.secret_key = 'replace_this_with_a_random_secret_key_12345'

DATABASE = 'fda_data.db'

# Function to get a database connection
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# Create comprehensive database tables
def init_db():
    with get_db_connection() as conn:
        # Main events table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                report_number TEXT,
                event_type TEXT,
                event_location TEXT,
                date_received TEXT,
                date_of_event TEXT,
                report_date TEXT,
                date_facility_aware TEXT,
                date_added TEXT,
                date_changed TEXT,
                report_to_fda TEXT,
                adverse_event_flag TEXT,
                product_problem_flag TEXT,
                report_source_code TEXT,
                health_professional TEXT,
                number_devices_in_event TEXT,
                number_patients_in_event TEXT,
                noe_summarized TEXT,
                manufacturer_name TEXT,
                manufacturer_address_1 TEXT,
                manufacturer_address_2 TEXT,
                manufacturer_city TEXT,
                manufacturer_state TEXT,
                manufacturer_zip_code TEXT,
                manufacturer_country TEXT,
                manufacturer_postal_code TEXT,
                type_of_report TEXT,
                remedial_action TEXT,
                raw_json TEXT
            )
        ''')
        
        # Device details table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS devices (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                device_sequence_number TEXT,
                brand_name TEXT,
                generic_name TEXT,
                manufacturer_d_name TEXT,
                model_number TEXT,
                catalog_number TEXT,
                lot_number TEXT,
                device_operator TEXT,
                device_availability TEXT,
                device_report_product_code TEXT,
                device_name TEXT,
                medical_specialty_description TEXT,
                regulation_number TEXT,
                device_class TEXT,
                implant_flag TEXT,
                raw_device_json TEXT,
                FOREIGN KEY (event_id) REFERENCES events (id)
            )
        ''')
        
        # Patient details table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS patients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                patient_sequence_number TEXT,
                patient_age TEXT,
                patient_sex TEXT,
                patient_weight TEXT,
                patient_ethnicity TEXT,
                patient_race TEXT,
                sequence_number_outcome TEXT,
                sequence_number_treatment TEXT,
                raw_patient_json TEXT,
                FOREIGN KEY (event_id) REFERENCES events (id)
            )
        ''')
        
        # MDR text table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS mdr_texts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_id INTEGER,
                text_type_code TEXT,
                patient_sequence_number TEXT,
                text TEXT,
                mdr_text_key TEXT,
                FOREIGN KEY (event_id) REFERENCES events (id)
            )
        ''')
        
        conn.commit()

# Enhanced fetch function with pagination
def fetch_all_API_data(base_query, max_records=None):
    all_data = []
    skip = 0
    limit = 1000  # Maximum allowed by FDA API
    
    while True:
        query = f"{base_query}&limit={limit}&skip={skip}"
        print(f"Fetching records {skip} to {skip + limit}...")
        
        response = requests.get(query)
        if response.status_code != 200:
            print(f"Error fetching data: {response.status_code}")
            break
            
        data = response.json()
        results = data.get('results', [])
        
        if not results:
            break
            
        all_data.extend(results)
        
        # Check if we've reached the maximum or if there are no more results
        total_results = data.get('meta', {}).get('results', {}).get('total', 0)
        if skip + limit >= total_results or (max_records and len(all_data) >= max_records):
            break
            
        skip += limit
    
    return all_data

# Enhanced save function for comprehensive data
def save_comprehensive_data(data):
    with get_db_connection() as conn:
        for record in data:
            # Insert main event record
            cursor = conn.execute('''
                INSERT INTO events (
                    report_number, event_type, event_location, date_received, 
                    date_of_event, report_date, date_facility_aware, date_added,
                    date_changed, report_to_fda, adverse_event_flag, product_problem_flag,
                    report_source_code, health_professional, number_devices_in_event,
                    number_patients_in_event, noe_summarized, manufacturer_name,
                    manufacturer_address_1, manufacturer_address_2, manufacturer_city,
                    manufacturer_state, manufacturer_zip_code, manufacturer_country,
                    manufacturer_postal_code, type_of_report, remedial_action, raw_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                record.get('report_number'),
                record.get('event_type'),
                record.get('event_location'),
                record.get('date_received'),
                record.get('date_of_event'),
                record.get('report_date'),
                record.get('date_facility_aware'),
                record.get('date_added'),
                record.get('date_changed'),
                record.get('report_to_fda'),
                record.get('adverse_event_flag'),
                record.get('product_problem_flag'),
                record.get('report_source_code'),
                record.get('health_professional'),
                record.get('number_devices_in_event'),
                record.get('number_patients_in_event'),
                record.get('noe_summarized'),
                record.get('manufacturer_name'),
                record.get('manufacturer_address_1'),
                record.get('manufacturer_address_2'),
                record.get('manufacturer_city'),
                record.get('manufacturer_state'),
                record.get('manufacturer_zip_code'),
                record.get('manufacturer_country'),
                record.get('manufacturer_postal_code'),
                json.dumps(record.get('type_of_report', [])),
                json.dumps(record.get('remedial_action', [])),
                json.dumps(record)
            ))
            
            event_id = cursor.lastrowid
            
            # Insert device records
            devices = record.get('device', [])
            for device in devices:
                openfda = device.get('openfda', {})
                conn.execute('''
                    INSERT INTO devices (
                        event_id, device_sequence_number, brand_name, generic_name,
                        manufacturer_d_name, model_number, catalog_number, lot_number,
                        device_operator, device_availability, device_report_product_code,
                        device_name, medical_specialty_description, regulation_number,
                        device_class, implant_flag, raw_device_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    event_id,
                    device.get('device_sequence_number'),
                    device.get('brand_name'),
                    device.get('generic_name'),
                    device.get('manufacturer_d_name'),
                    device.get('model_number'),
                    device.get('catalog_number'),
                    device.get('lot_number'),
                    device.get('device_operator'),
                    device.get('device_availability'),
                    device.get('device_report_product_code'),
                    openfda.get('device_name'),
                    openfda.get('medical_specialty_description'),
                    openfda.get('regulation_number'),
                    openfda.get('device_class'),
                    device.get('implant_flag'),
                    json.dumps(device)
                ))
            
            # Insert patient records
            patients = record.get('patient', [])
            for patient in patients:
                conn.execute('''
                    INSERT INTO patients (
                        event_id, patient_sequence_number, patient_age, patient_sex,
                        patient_weight, patient_ethnicity, patient_race,
                        sequence_number_outcome, sequence_number_treatment, raw_patient_json
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    event_id,
                    patient.get('patient_sequence_number'),
                    patient.get('patient_age'),
                    patient.get('patient_sex'),
                    patient.get('patient_weight'),
                    patient.get('patient_ethnicity'),
                    patient.get('patient_race'),
                    json.dumps(patient.get('sequence_number_outcome', [])),
                    json.dumps(patient.get('sequence_number_treatment', [])),
                    json.dumps(patient)
                ))
            
            # Insert MDR text records
            mdr_texts = record.get('mdr_text', [])
            for mdr_text in mdr_texts:
                conn.execute('''
                    INSERT INTO mdr_texts (
                        event_id, text_type_code, patient_sequence_number, text, mdr_text_key
                    ) VALUES (?, ?, ?, ?, ?)
                ''', (
                    event_id,
                    mdr_text.get('text_type_code'),
                    mdr_text.get('patient_sequence_number'),
                    mdr_text.get('text'),
                    mdr_text.get('mdr_text_key', '')
                ))
        
        conn.commit()



def sanitize_text(text):
    if not isinstance(text, str):
        return text
    # Remove all ASCII control characters except tab, newline, carriage return
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)
    # Remove invalid XML characters (0x00-0x08, 0x0B, 0x0C, 0x0E-0x1F, 0x7F-0x9F)
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]', '', text)
    # Replace problematic Unicode quotes with standard quotes
    text = text.replace('\u201c', '"').replace('\u201d', '"').replace('\u2018', "'").replace('\u2019', "'")
    # Replace other problematic Unicode characters
    text = text.replace('\u2013', '-').replace('\u2014', '-')  # en-dash, em-dash
    text = text.replace('\u2022', '•')  # bullet point
    text = text.replace('\u2026', '...')  # ellipsis
    # Remove any remaining control characters
    text = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', text)
    # Limit string length to prevent Excel issues (Excel has cell size limits)
    if len(text) > 32000:  # Excel cell limit is 32,767 characters
        text = text[:32000] + "..."
    return text

def sanitize_df(df):
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].apply(sanitize_text)
    return df



def extract_event_fields(event, field_list, event_id=None):
    """
    Extracts the specified fields from the event JSON, handling arrays and nested fields.
    For array fields, splits into _1, _2, ... columns, placed after the main field.
    For nested fields (device.*, patient.*, mdr_text.*), pivots as needed.
    """
    result = {}
    if event_id is not None:
        result['event_id'] = event_id
    # First, handle top-level fields and arrays
    for field in field_list:
        if '.' not in field and not field.startswith('device') and not field.startswith('patient') and not field.startswith('mdr_text'):
            value = event.get(field, '')
            if isinstance(value, list):
                result[field] = ''
                for i, v in enumerate(value):
                    result[f'{field}_{i+1}'] = v
            else:
                result[field] = value
    # Handle device fields
    devices = event.get('device', [])
    for i, device in enumerate(devices):
        for field in field_list:
            if field.startswith('device.'):
                subfield = field.split('.', 1)[1]
                value = device.get(subfield, '')
                if isinstance(value, list):
                    result[f'device_{subfield}_{i+1}'] = ''
                    for j, v in enumerate(value):
                        result[f'device_{subfield}_{i+1}_{j+1}'] = v
                else:
                    result[f'device_{subfield}_{i+1}'] = value
    # Handle patient fields
    patients = event.get('patient', [])
    for i, patient in enumerate(patients):
        for field in field_list:
            if field.startswith('patient.'):
                subfield = field.split('.', 1)[1]
                value = patient.get(subfield, '')
                if isinstance(value, list):
                    result[f'patient_{subfield}_{i+1}'] = ''
                    for j, v in enumerate(value):
                        result[f'patient_{subfield}_{i+1}_{j+1}'] = v
                else:
                    result[f'patient_{subfield}_{i+1}'] = value
    # Handle mdr_text fields (not in Events sheet, but for completeness)
    # MAUDE report link
    mdr_report_key = event.get('mdr_report_key', '')
    maude_link = ''
    if mdr_report_key:
        # Try to get first device's product code and sequence number
        pc = ''
        seq = ''
        if devices and isinstance(devices, list):
            pc = devices[0].get('device_report_product_code', '')
            seq = devices[0].get('device_sequence_number', '')
        if pc and seq:
            maude_link = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmaude/detail.cfm?mdrfoi__id={mdr_report_key}&pc={pc}&device_sequence_no={seq}"
        else:
            maude_link = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmaude/detail.cfm?mdrfoi__id={mdr_report_key}"
    result['maude_report_link'] = maude_link
    return result

def export_to_excel(include_raw_events=True):
    import pandas as pd
    import json
    import os
    from datetime import datetime
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    fields_path = os.path.join(BASE_DIR, 'fields.xlsx')
    # Main fields and array fields for the Main Fields sheet
    main_fields = [
        'event_id', 'report_number', 'mdr_report_key', 'maude_report_link',
        'date_of_event', 'date_report', 'date_received', 'date_manufacturer_received',
        'device_date_received', 'device_expiration_date_of_device', 'patient_date_received',
        'device_generic_name', 'device_brand_name', 'device_manufacturer_d_name',
        'device_device_report_product_code', 'device_model_number', 'device_lot_number',
        'device_device_availability', 'device_device_evaluated_by_manufacturer', 'device_manufacturer_d_country',
        'single_use_flag', 'reprocessed_and_reused_flag', 'device_device_operator', 'report_source_code',
        'health_professional', 'reporter_occupation_code', 'source_type',
        'patient_patient_age', 'patient_patient_sex', 'patient_patient_weight', 'patient_patient_ethnicity', 'patient_patient_race',
        'event_type', 'adverse_event_flag', 'patient_patient_problems', 'patient_sequence_number_outcome', 'patient_sequence_number_treatment',
        'product_problem_flag', 'product_problems'
    ]
    array_fields = [
        'product_problems', 'patient_patient_problems', 'patient_sequence_number_outcome', 'patient_sequence_number_treatment'
    ]
    
    # Use the user-provided field list (excluding distributor and manufacturer contact fields for performance)
    field_list = [
        'adverse_event_flag', 'product_problems', 'product_problem_flag', 'date_of_event', 'date_report', 'date_received',
        'device_date_of_manufacturer', 'event_type', 'number_devices_in_event', 'number_patients_in_event', 'previous_use_code',
        'remedial_action', 'removal_correction_number', 'report_number', 'single_use_flag', 'report_source_code',
        'health_professional', 'reporter_occupation_code', 'initial_report_to_fda', 'reprocessed_and_reused_flag',
        'device.device_event_key', 'device.date_received', 'device.brand_name',
        'device.generic_name', 'device.device_report_product_code',
        'device.model_number', 'device.catalog_number', 'device.lot_number', 'device.other_id_number',
        'device.expiration_date_of_device', 'device.device_availability',
        'device.device_evaluated_by_manufacturer', 'device.device_operator',
        'device.implant_flag', 'device.date_removed_flag', 'device.manufacturer_d_name',
        'device.manufacturer_d_country', 'device.device_class', 'device.device_name', 'device.fei_number', 'device.medical_specialty_description', 'device.registration_number', 'patient.date_received', 'patient.patient_age',
        'patient.patient_sex', 'patient.patient_weight', 'patient.patient_ethnicity', 'patient.patient_race',
        'patient.patient_problems', 'patient.sequence_number_outcome', 'patient.sequence_number_treatment',
        'mdr_text.date_report', 'mdr_text.mdr_text_key', 'mdr_text.patient_sequence_number', 'mdr_text.text',
        'mdr_text.text_type_code', 'type_of_report', 'date_facility_aware', 'report_date', 'report_to_fda',
        'date_report_to_fda', 'report_to_manufacturer', 'date_report_to_manufacturer', 'event_location', 'manufacturer_name', 'manufacturer_address_1', 'manufacturer_address_2',
        'manufacturer_city', 'manufacturer_country', 'manufacturer_gl_name',
        'manufacturer_gl_country', 'date_manufacturer_received',
        'source_type', 'event_key', 'mdr_report_key', 'device name', 'fei_number',
        'medical_specialty_description', 'registration_number', 'regulation_number'
    ]
    # List of array fields to handle placement
    array_fields = [
        'product_problems', 'remedial_action', 'previous_use_code', 'removal_correction_number', 'single_use_flag',
        'report_source_code', 'reporter_occupation_code', 'initial_report_to_fda', 'reprocessed_and_reused_flag',
        'patient.patient_problems', 'patient.sequence_number_outcome', 'patient.sequence_number_treatment'
    ]
    def format_all_date_columns(df):
        import re
        from datetime import datetime
        date_pattern = re.compile(r'^\d{8}$')
        date_base_names = [
            'date_of_event', 'date_report', 'date_received', 'date_manufacturer_received',
            'device_date_received', 'device_expiration_date_of_device', 'patient_date_received',
            'device_date_of_manufacturer', 'device.date_received', 'device.expiration_date_of_device',
            'device.date_returned_to_manufacturer', 'device.date_removed_flag', 'mdr_text.date_report',
            'date_facility_aware', 'report_date', 'date_report_to_fda', 'date_report_to_manufacturer'
        ]
        date_cols_to_format = [col for col in df.columns if any(col == base or col.startswith(base + '_') for base in date_base_names)]
        for col in date_cols_to_format:
            df[col] = df[col].apply(
                lambda x: datetime.strptime(str(x), '%Y%m%d').strftime('%m/%d/%Y') if isinstance(x, str) and date_pattern.match(x) else x
            )
        return df

    def translate_fda_codes(df):
        """Translate FDA codes to human-readable text for user-friendly display"""
        
        # Patient Outcome Codes (based on actual data analysis)
        outcome_mapping = {
            'R': 'Required Intervention',
            'O': 'Other',
            'H': 'Hospitalization',
            'D': 'Death',
            'L': 'Life Threatening',
            'I': 'Injury',
            'M': 'Malfunction',
            'N': 'No Information',
            'U': 'Unknown',
            'S': 'Disability'
        }
        
        # Device Evaluated by Manufacturer Codes (based on actual data)
        device_evaluated_mapping = {
            'R': 'Returned to Manufacturer',
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Reporter Occupation Codes (based on actual data)
        occupation_mapping = {
            '501': 'Administrator/Supervisor',
            '003': 'Non-Healthcare Professional',
            '117': 'Nurse Practitioner',
            '2': 'Nurse',
            'PHYSICIAN': 'Physician',
            'NURSE': 'Nurse',
            'OTHER': 'Other',
            'OTHER HEALTH CARE PROFESSIONAL': 'Other Health Care Professional',
            'RISK MANAGER': 'Risk Manager',
            'PATIENT': 'Patient',
            'ATTORNEY': 'Attorney',
            'PATIENT FAMILY MEMBER OR FRIEND': 'Patient Family Member or Friend',
            'UNKNOWN': 'Unknown'
        }
        
        # Previous Use Codes (based on actual data)
        previous_use_mapping = {
            'I': 'Invalid/Incomplete',
            'N': 'No',
            'U': 'Unknown',
            '*': 'Not Available'
        }
        
        # Report to FDA Codes (based on actual data)
        report_to_fda_mapping = {
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Health Professional Codes (based on actual data)
        health_prof_mapping = {
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Single Use Flag (based on actual data)
        single_use_mapping = {
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Reprocessed Flag (based on actual data)
        reprocessed_mapping = {
            'N': 'No',
            'I': 'Invalid/Incomplete'
        }
        
        # Adverse Event Flag (based on actual data)
        adverse_event_mapping = {
            'Y': 'Yes',
            'N': 'No'
        }
        
        # Product Problem Flag (based on actual data)
        product_problem_mapping = {
            'Y': 'Yes',
            'N': 'No',
            '*': 'Not Available'
        }
        
        # Device Operator Codes (based on actual data)
        device_operator_mapping = {
            'I': 'No Information',
            '0': 'Other',
            'HEALTH PROFESSIONAL': 'Health Professional',
            'LAY USER/PATIENT': 'Lay User/Patient',
            'INVALID DATA': 'Invalid Data',
            'PHYSICIAN': 'Physician',
            'OTHER': 'Other'
        }
        
        # Event Location Codes (based on actual data)
        event_location_mapping = {
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Manufacturer Link Flag Codes (based on actual data)
        manufacturer_link_flag_mapping = {
            'Y': 'Yes',
            'N': 'No',
            'I': 'Invalid/Incomplete',
            '*': 'Not Available'
        }
        
        # Apply translations to relevant columns
        for col in df.columns:
            if 'Patient Outcome' in col:
                # Handle semicolon-separated codes for Patient Outcome
                def translate_outcome_codes(value):
                    if pd.isna(value) or value == '':
                        return value
                    if isinstance(value, str):
                        # Split by semicolon and translate each code
                        codes = [code.strip() for code in value.split(';')]
                        translated_codes = []
                        for code in codes:
                            if code in outcome_mapping:
                                translated_codes.append(outcome_mapping[code])
                            else:
                                translated_codes.append(code)  # Keep original if not in mapping
                        return '; '.join(translated_codes)
                    return value
                df[col] = df[col].apply(translate_outcome_codes)
            elif 'Device Evaluated By Manufacturer' in col:
                df[col] = df[col].map(device_evaluated_mapping).fillna(df[col])
            elif 'Reporter Occupation Code' in col:
                df[col] = df[col].map(occupation_mapping).fillna(df[col])
            elif 'Previous Use Code' in col:
                df[col] = df[col].map(previous_use_mapping).fillna(df[col])
            elif 'Report To FDA' in col:
                df[col] = df[col].map(report_to_fda_mapping).fillna(df[col])
            elif 'Health Professional' in col:
                df[col] = df[col].map(health_prof_mapping).fillna(df[col])
            elif 'Single Use Flag' in col:
                df[col] = df[col].map(single_use_mapping).fillna(df[col])
            elif 'Reprocessed And Reused Flag' in col:
                df[col] = df[col].map(reprocessed_mapping).fillna(df[col])
            elif 'Adverse Event Flag' in col:
                df[col] = df[col].map(adverse_event_mapping).fillna(df[col])
            elif 'Product Problem Flag' in col:
                df[col] = df[col].map(product_problem_mapping).fillna(df[col])
            elif 'Device Operator' in col:
                df[col] = df[col].map(device_operator_mapping).fillna(df[col])
            elif 'Event Location' in col:
                df[col] = df[col].map(event_location_mapping).fillna(df[col])
            elif 'Manufacturer Link Flag' in col:
                df[col] = df[col].map(manufacturer_link_flag_mapping).fillna(df[col])
        
        return df

    with get_db_connection() as conn:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        filename = f'MAUDEMetrics_{timestamp}.xlsx'
        mdr_texts_csv = f'fda_mdr_texts_full_{timestamp}.csv'
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                import re
                def extract_numeric(val):
                    if pd.isnull(val):
                        return None
                    match = re.search(r'\d+(\.\d+)?', str(val))
                    return float(match.group()) if match else None
                

                
                # 1. EVENTS - Extract only user-specified fields, event_id as first column (OPTIMIZED)
                events_query = 'SELECT id, raw_json FROM events ORDER BY id'
                events_df = pd.read_sql_query(events_query, conn)
                print('Events DataFrame shape:', events_df.shape)
                
                # OPTIMIZATION 1: Pre-compile regex patterns for better performance
                import re
                date_pattern = re.compile(r'^\d{8}$')
                age_pattern = re.compile(r'\s*(YR|YEARS|YEAR|YRS)\s*', re.IGNORECASE)
                number_pattern = re.compile(r'\d+')
                
                # OPTIMIZATION 2: Use list comprehension instead of append for better performance
                all_events_data = []
                for _, row in events_df.iterrows():
                    if row['raw_json']:
                        try:
                            event = json.loads(row['raw_json'])
                            extracted = extract_event_fields(event, field_list, event_id=row['id'])
                            all_events_data.append(extracted)
                        except Exception as e:
                            all_events_data.append({'event_id': row['id'], 'error': str(e)})
                
                if not all_events_data:
                    raise Exception('No data to export. Please run a search and try again.')
                
                if all_events_data:
                    events_flat_df = pd.DataFrame(all_events_data)
                    events_flat_df = sanitize_df(events_flat_df)
                    events_flat_df = format_all_date_columns(events_flat_df)
                    # Build ordered columns: event_id, then for each field, its array columns immediately after
                    ordered_cols = ['event_id']
                    for field in field_list:
                        if field in events_flat_df.columns:
                            ordered_cols.append(field)
                        # Add array columns immediately after
                        i = 1
                        while f'{field}_{i}' in events_flat_df.columns:
                            ordered_cols.append(f'{field}_{i}')
                            i += 1
                        # Insert maude_report_link after mdr_report_key
                        if field == 'mdr_report_key' and 'maude_report_link' in events_flat_df.columns:
                            ordered_cols.append('maude_report_link')
                    # Add maude_report_link at the end if not already added
                    if 'maude_report_link' in events_flat_df.columns and 'maude_report_link' not in ordered_cols:
                        ordered_cols.append('maude_report_link')
                    # Add any extra columns
                    extra_cols = [col for col in events_flat_df.columns if col not in ordered_cols]
                    events_flat_df = events_flat_df[ordered_cols + extra_cols]
                    
                    # Only write Raw_Events sheet if requested
                    if include_raw_events:
                        print("Writing Raw_Events sheet to Excel...")
                        events_flat_df.to_excel(writer, sheet_name='Raw_Events', index=False)
                    else:
                        print("Skipping Raw_Events sheet for better performance...")
                    # --- Restore main_fields_df construction ---
                    main_fields_cols = []
                    for field in main_fields:
                        # Find all columns in events_flat_df that start with this field name (robust: any suffix)
                        matching_cols = [col for col in events_flat_df.columns if col == field or col.startswith(field + '_')]
                        main_fields_cols.extend(matching_cols)
                    # Add any extra columns not in main_fields_cols
                    extra_cols = [col for col in events_flat_df.columns if col not in main_fields_cols]
                    main_fields_df = events_flat_df[main_fields_cols + extra_cols]
                    main_fields_df = format_all_date_columns(main_fields_df)
                    
                    # Remove completely blank columns (no data except header)
                    blank_cols = []
                    for col in main_fields_df.columns:
                        # Check if the column has any non-blank data (excluding header)
                        has_data = False
                        for val in main_fields_df[col]:
                            # Check for actual data (not NaN, None, empty string, or whitespace)
                            if pd.notna(val) and val is not None and str(val).strip() != '':
                                has_data = True
                                break
                        if not has_data:
                            blank_cols.append(col)
                    
                    # Remove columns with no data
                    if blank_cols:
                        print(f"Removing {len(blank_cols)} blank columns from Custom_Events: {blank_cols}")
                        main_fields_df = main_fields_df.drop(columns=blank_cols)
                    
                    # Enhanced humanize function for professional column naming
                    def enhanced_humanize(col):
                        if not isinstance(col, str):
                            return col
                        field_mapping = {
                            'event_id': 'Event ID',
                            'report_number': 'Report Number',
                            'mdr_report_key': 'MDR Report Key',
                            'maude_report_link': 'MAUDE Report Link',
                            'date_of_event': 'Event Date',
                            'date_report': 'Report Date',
                            'date_received': 'Date Received',
                            'date_manufacturer_received': 'Date Manufacturer Received',
                            'device_date_received': 'Device Date Received',
                            'device_expiration_date_of_device': 'Device Expiration Date',
                            'patient_date_received': 'Patient Date Received',
                            'device_generic_name': 'Product Class',
                            'device_brand_name': 'Brand Name',
                            'device_manufacturer_d_name': 'Manufacturer',
                            'device_device_report_product_code': 'Product Code',
                            'device_model_number': 'Model Number',
                            'device_catalog_number': 'Catalog Number',
                            'device_lot_number': 'Lot Number',
                            'device_device_availability': 'Device Availability',
                            'device_device_evaluated_by_manufacturer': 'Device Evaluated By Manufacturer',
                            'device_manufacturer_d_country': 'Manufacturer Country',
                            'single_use_flag': 'Single Use Flag',
                            'reprocessed_and_reused_flag': 'Reprocessed And Reused Flag',
                            'device_device_operator': 'Device Operator',
                            'report_source_code': 'Report Source Code',
                            'health_professional': 'Health Professional',
                            'reporter_occupation_code': 'Reporter Occupation Code',
                            'source_type': 'Source Type',
                            'patient_patient_age': 'Patient Age',
                            'patient_patient_sex': 'Patient Sex',
                            'patient_patient_weight': 'Patient Weight',
                            'patient_patient_ethnicity': 'Patient Ethnicity',
                            'patient_patient_race': 'Patient Race',
                            'event_type': 'Event Type',
                            'adverse_event_flag': 'Adverse Event Flag',
                            'patient_patient_problems': 'Patient Problem',
                            'patient_sequence_number_outcome': 'Patient Outcome',
                            'patient_sequence_number_treatment': 'Patient Treatment',
                            'product_problem_flag': 'Product Problem Flag',
                            'product_problems': 'Device Problem',
                            'date_report_to_fda': 'Date Report To FDA',
                            'date_report_to_manufacturer': 'Date Report To Manufacturer'
                        }
                        if col in field_mapping:
                            return field_mapping[col]
                        # Handle array fields with numbers
                        if '_' in col and col[-1].isdigit():
                            base_field = col.rsplit('_', 1)[0]
                            number = col.rsplit('_', 1)[1]
                            if base_field in field_mapping:
                                return f"{field_mapping[base_field]} {number}"
                            # Handle special cases for nested array fields - clean up the naming
                            elif base_field.endswith('_1') and base_field[:-2] in field_mapping:
                                base_base_field = base_field[:-2]
                                # Clean up nested array naming (e.g., "Patient Patient Problems 1 1" -> "Patient Problem 1")
                                if base_base_field == 'patient_patient_problems':
                                    return f"Patient Problem {number}"
                                elif base_base_field == 'patient_sequence_number_outcome':
                                    return f"Patient Outcome {number}"
                                elif base_base_field == 'patient_sequence_number_treatment':
                                    return f"Patient Treatment {number}"
                                else:
                                    return f"{field_mapping[base_base_field]} {number}"
                        return col.replace('_', ' ').replace('.', ' ').title()
                    
                    # Apply enhanced column naming
                    main_fields_df.columns = [enhanced_humanize(c) for c in main_fields_df.columns]
                    
                    # Intelligent column reordering for Custom_Events sheet
                    # Define base priority order (without array numbers)
                    base_priority_order = [
                        'Event ID', 'Report Number', 'MDR Report Key', 'MAUDE Report Link', 'Event Date', 'Report Date', 'Date Received',
                        'Date Report To FDA', 'Date Report To Manufacturer', 'Date Manufacturer Received', 'Device Date Received', 'Device Expiration Date', 'Patient Date Received',
                        'Product Class', 'Brand Name', 'Product Code', 'Model Number', 'Manufacturer', 'Manufacturer Country', 'Lot Number', 'Catalog Number', 'Device Availability', 'Device Evaluated By Manufacturer', 'Single Use Flag', 'Reprocessed And Reused Flag', 'Device Operator', 'Report Source Code', 'Health Professional', 'Reporter Occupation Code', 'Source Type', 'Patient Age', 'Patient Sex', 'Patient Weight', 'Patient Ethnicity', 'Patient Race', 'Event Type',
                        'Adverse Event Flag', 'Product Problem Flag', 'Device Problem', 'Patient Problem', 'Patient Outcome', 'Patient Treatment'
                    ]
                    
                    # Build complete priority columns maintaining original position order
                    priority_columns = []
                    for base_field in base_priority_order:
                        # Add the base field first (if it exists in the dataframe)
                        if base_field in main_fields_df.columns:
                            priority_columns.append(base_field)
                        
                        # Add all numbered variations of this field in sequence
                        for i in range(1, 50):  # Support up to 49 array elements
                            numbered_field = f"{base_field} {i}"
                            if numbered_field in main_fields_df.columns:
                                priority_columns.append(numbered_field)
                        
                        # Handle nested array fields (e.g., "Patient Problems 1 1", "Patient Problems 1 2", etc.)
                        # These come from nested structures like patient.patient_problems
                        for i in range(1, 50):  # First level array
                            for j in range(1, 50):  # Second level array
                                nested_field = f"{base_field} {i} {j}"
                                if nested_field in main_fields_df.columns:
                                    priority_columns.append(nested_field)
                        
                        # Continue to next base field (arrays maintain their position in sequence)
                    
                    # Reorder columns: priority columns first, then others
                    available_priority_cols = [col for col in priority_columns if col in main_fields_df.columns]
                    other_cols = [col for col in main_fields_df.columns if col not in available_priority_cols]
                    reordered_cols = available_priority_cols + other_cols
                    
                    # Apply reordering
                    main_fields_df = main_fields_df[reordered_cols]
                    
                    # FDA Code Analysis removed for cleaner output
                    
                    # Apply FDA code translation for user-friendly display
                    main_fields_df = translate_fda_codes(main_fields_df)
                    
                    main_fields_df.to_excel(writer, sheet_name='Events', index=False)
                # 2. MDR TEXTS - Add event_id, mdr_report_key, and maude_report_link (link only for first row per event_id)
                mdr_texts_query = 'SELECT * FROM mdr_texts ORDER BY event_id, text_type_code'
                mdr_texts_df = pd.read_sql_query(mdr_texts_query, conn)
                # OPTIMIZATION 5: Reuse already processed event data instead of re-parsing JSON
                event_keys = {}
                event_links = {}
                for _, row in events_df.iterrows():
                    try:
                        event = json.loads(row['raw_json'])
                        event_keys[row['id']] = event.get('mdr_report_key', '')
                        mdr_report_key = event.get('mdr_report_key', '')
                        devices = event.get('device', [])
                        maude_link = ''
                        if mdr_report_key:
                            pc = ''
                            seq = ''
                            if devices and isinstance(devices, list) and devices:
                                pc = devices[0].get('device_report_product_code', '')
                                seq = devices[0].get('device_sequence_number', '')
                            if pc and seq:
                                maude_link = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmaude/detail.cfm?mdrfoi__id={mdr_report_key}&pc={pc}&device_sequence_no={seq}"
                            else:
                                maude_link = f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfmaude/detail.cfm?mdrfoi__id={mdr_report_key}"
                        event_links[row['id']] = maude_link
                    except:
                        event_keys[row['id']] = ''
                        event_links[row['id']] = ''
                
                # Use mdr_text_key from the mdr_texts table instead of mdr_report_key from events
                # The mdr_text_key column is already in the mdr_texts_df from the database query
                
                # OPTIMIZATION: Use vectorized operations for maude_report_link assignment
                mdr_texts_df['maude_report_link'] = mdr_texts_df['event_id'].map(event_links)
                mdr_texts_df = sanitize_df(mdr_texts_df)
                mdr_texts_df = format_all_date_columns(mdr_texts_df)
                # Reorder and filter columns for MDR_Texts sheet
                mdr_cols = ['event_id', 'maude_report_link', 'text_type_code', 'text']
                mdr_texts_df = mdr_texts_df[[col for col in mdr_cols if col in mdr_texts_df.columns]]
                
                # Remove repeated values in MDR_Texts sheet (keep first occurrence, empty the rest)
                dedup_columns = ['event_id', 'maude_report_link', 'text_type_code']
                for col in dedup_columns:
                    if col in mdr_texts_df.columns:
                        prev_value = None
                        for idx in mdr_texts_df.index:
                            current_value = mdr_texts_df.at[idx, col]
                            if current_value == prev_value:
                                mdr_texts_df.at[idx, col] = ''
                            else:
                                prev_value = current_value
                
                mdr_texts_df.to_excel(writer, sheet_name='MDR_Texts', index=False)
                # --- Improved All-in-One Summary Sheet ---
                summary_blocks = []
                # 1. Total Reports
                total_reports = len(events_flat_df)
                summary_blocks.append(pd.DataFrame({
                    'Summary': ['Total Reports'],
                    'Value': [total_reports]
                }))
                summary_blocks.append(pd.DataFrame({'': ['']}))
                
                # 2. Patient Demographics Table (improved formatting)
                demo_table = []
                # OPTIMIZATION 6: Optimized demographic calculations
                # Age
                age_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_age')]
                if age_cols:
                    ages = pd.Series([extract_numeric(v) for v in events_flat_df[age_cols].values.flatten()]).dropna()
                    age_val = f"{int(ages.median())} ({int(ages.min())}-{int(ages.max())})" if not ages.empty else "N/A"
                    demo_table.append(["Age (years) median (range)", age_val, "", ""])
                # Weight
                weight_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_weight')]
                if weight_cols:
                    weights = pd.Series([extract_numeric(v) for v in events_flat_df[weight_cols].values.flatten()]).dropna()
                    weight_val = f"{weights.median():.1f} ({weights.min():.1f}-{weights.max():.1f})" if not weights.empty else "N/A"
                    demo_table.append(["Weight median (range)", weight_val, "", ""])
                # Sex
                sex_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_sex')]
                sex_vals = pd.Series(events_flat_df[sex_cols].values.flatten()).dropna()
                if not sex_vals.empty:
                    first = True
                    for k, v in sex_vals.value_counts().items():
                        if first:
                            demo_table.append(["Sex", k, v, f"{100*v/len(sex_vals):.1f}%"])
                            first = False
                        else:
                            demo_table.append(["", k, v, f"{100*v/len(sex_vals):.1f}%"])
                # Ethnicity
                eth_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_ethnicity')]
                eth_vals = pd.Series(events_flat_df[eth_cols].values.flatten()).dropna()
                if not eth_vals.empty:
                    first = True
                    for k, v in eth_vals.value_counts().items():
                        if first:
                            demo_table.append(["Ethnicity", k, v, f"{100*v/len(eth_vals):.1f}%"])
                            first = False
                        else:
                            demo_table.append(["", k, v, f"{100*v/len(eth_vals):.1f}%"])
                # Race
                race_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_race')]
                race_vals = pd.Series(events_flat_df[race_cols].values.flatten()).dropna()
                if not race_vals.empty:
                    first = True
                    for k, v in race_vals.value_counts().items():
                        if first:
                            demo_table.append(["Race", k, v, f"{100*v/len(race_vals):.1f}%"])
                            first = False
                        else:
                            demo_table.append(["", k, v, f"{100*v/len(race_vals):.1f}%"])
                demo_df = pd.DataFrame(demo_table, columns=["Patient Demographics", "Value", "Frequency", "Percentage"])
                summary_blocks.append(demo_df)
                summary_blocks.append(pd.DataFrame({'': ['']}))
                # 3. Event/Product Characteristics: each table with column headers, no section header, no blank rows between
                event_fields = [
                    ('event_type', 'Event Type'),
                    ('report_source_code', 'Report Source Code'),
                    ('source_type', 'Source Type'),
                    ('reporter_occupation_code', 'Reporter Occupation Code'),
                    ('device_device_report_product_code', 'Product Code'),
                    ('device_model_number', 'Model Number'),
                    ('device_manufacturer_d_name', 'Manufacturer'),
                    ('device_manufacturer_d_country', 'Manufacturer Country'),
                    ('device_brand_name', 'Brand Name'),
                    ('device_generic_name', 'Product Class')
                ]
                for field, label in event_fields:
                    cols = [c for c in events_flat_df.columns if c.startswith(field)]
                    vals = pd.Series(events_flat_df[cols].values.flatten()).dropna()
                    if not vals.empty:
                        counts = vals.value_counts()
                        table_rows = []
                        table_rows.append([label, "Frequency", "Percentage"])
                        for v in counts.index:
                            table_rows.append([v, counts[v], f"{100*counts[v]/len(vals):.1f}%"])
                        event_df = pd.DataFrame(table_rows[1:], columns=table_rows[0])
                        summary_blocks.append(event_df)
                summary_blocks.append(pd.DataFrame({'': ['']}))
                # 4. Device Problem Table (no table header)
                prod_cols = [c for c in events_flat_df.columns if c.startswith('product_problems')]
                prod_vals = pd.Series(events_flat_df[prod_cols].values.flatten()).dropna()
                prod_counts = prod_vals.value_counts()
                prod_table_start = None
                if not prod_counts.empty:
                    prod_df = pd.DataFrame({'Device Problem': prod_counts.index, 'Frequency': prod_counts.values})
                    prod_df['Percentage'] = prod_df['Frequency'].apply(lambda v: f"{100*v/len(prod_vals):.1f}%" if len(prod_vals) else '0%')
                    prod_table_start = len(summary_blocks)
                    summary_blocks.append(prod_df)
                    summary_blocks.append(pd.DataFrame({'': ['']}))
                # 5. Patient Problem Table (no table header)
                pprob_cols = [c for c in events_flat_df.columns if c.startswith('patient_patient_problems')]
                pprob_vals = pd.Series(events_flat_df[pprob_cols].values.flatten()).dropna()
                pprob_counts = pprob_vals.value_counts()
                pprob_table_start = None
                if not pprob_counts.empty:
                    pprob_df = pd.DataFrame({'Patient Problem': pprob_counts.index, 'Frequency': pprob_counts.values})
                    pprob_df['Percentage'] = pprob_df['Frequency'].apply(lambda v: f"{100*v/len(pprob_vals):.1f}%" if len(pprob_vals) else '0%')
                    pprob_table_start = len(summary_blocks)
                    summary_blocks.append(pprob_df)
                    summary_blocks.append(pd.DataFrame({'': ['']}))
                # Write all summary blocks to the Summary sheet in the same writer session
                # Add Events Missing Patient Data at the end
                missing_patients = pd.read_sql_query('''
                    SELECT e.id as event_id, e.report_number
                    FROM events e
                    LEFT JOIN patients p ON e.id = p.event_id
                    WHERE p.id IS NULL
                ''', conn)
                if not missing_patients.empty:
                    summary_blocks.append(pd.DataFrame({'Summary': ['Events Missing Patient Data']}))
                    missing_patients_df = missing_patients.rename(columns={
                        'event_id': 'Event ID',
                        'report_number': 'Report Number'
                    })
                    summary_blocks.append(missing_patients_df)
                    summary_blocks.append(pd.DataFrame({'': ['']}))
                startrow = 0
                table_starts = []
                for block in summary_blocks:
                    block.columns = [enhanced_humanize(c) for c in block.columns]
                    table_starts.append(startrow)
                    block.to_excel(writer, sheet_name='Summary', index=False, startrow=startrow, header=True)
                    startrow += len(block) + 2
                # OPTIMIZATION 7: Simplified Excel formatting for better performance
                from openpyxl.styles import PatternFill, Font
                from openpyxl.chart import BarChart, Reference
                wb = writer.book
                ws = wb['Summary']
                # Color map for tables
                table_colors = ["34495E", "27AE60", "E67E22", "8E44AD", "2980B9"]
                color_idx = 0
                for i, start in enumerate(table_starts):
                    # OPTIMIZATION: Only apply formatting to non-empty cells
                    for row in ws.iter_rows(min_row=start+1, max_row=start+1):
                        for cell in row:
                            if cell.value:
                                cell.fill = PatternFill(start_color=table_colors[color_idx%len(table_colors)], end_color=table_colors[color_idx%len(table_colors)], fill_type='solid')
                                break  # Only format first non-empty cell in row
                    color_idx += 1
                # Add Excel bar charts for Device Problem and Patient Problem
                def find_table_data_range(ws, header):
                    # Find the header row
                    for row in ws.iter_rows():
                        if row[0].value == header:
                            header_row = row[0].row
                            break
                    else:
                        return None, None
                    # Data starts after header
                    data_start = header_row + 1
                    # Data ends at first blank in col A
                    data_end = data_start
                    while ws[f'A{data_end}'].value:
                        data_end += 1
                    return data_start, data_end-1
                # Device Problem chart
                prod_data_start, prod_data_end = find_table_data_range(ws, "Device Problem")
                if prod_data_start and prod_data_end > prod_data_start:
                    chart = BarChart()
                    chart.type = "bar"
                    chart.style = 10
                    chart.title = "Device Problem"
                    chart.y_axis.title = "Problem"
                    chart.x_axis.title = "Frequency"
                    data = Reference(ws, min_col=2, min_row=prod_data_start, max_col=2, max_row=prod_data_end)
                    cats = Reference(ws, min_col=1, min_row=prod_data_start, max_row=prod_data_end)
                    chart.add_data(data, titles_from_data=False)
                    chart.set_categories(cats)
                    chart.shape = 4
                    chart.height = max(7, (prod_data_end-prod_data_start+1)*0.5)
                    ws.add_chart(chart, f'E{prod_data_start}')
                # Patient Problem chart
                pprob_data_start, pprob_data_end = find_table_data_range(ws, "Patient Problem")
                if pprob_data_start and pprob_data_end > pprob_data_start:
                    chart = BarChart()
                    chart.type = "bar"
                    chart.style = 11
                    chart.title = "Patient Problem"
                    chart.y_axis.title = "Problem"
                    chart.x_axis.title = "Frequency"
                    data = Reference(ws, min_col=2, min_row=pprob_data_start, max_col=2, max_row=pprob_data_end)
                    cats = Reference(ws, min_col=1, min_row=pprob_data_start, max_row=pprob_data_end)
                    chart.add_data(data, titles_from_data=False)
                    chart.set_categories(cats)
                    chart.shape = 4
                    chart.height = max(7, (pprob_data_end-pprob_data_start+1)*0.5)
                    ws.add_chart(chart, f'E{pprob_data_start}')
                # Add Fields Reference sheet (fields.xlsx)
                if os.path.exists(fields_path):
                    fields_df = pd.read_excel(fields_path)
                    fields_df.to_excel(writer, sheet_name='Fields Reference', index=False)
        except Exception as e:
            print(f"Error in export: {str(e)}")
            raise e
        # --- Excel formatting with openpyxl ---
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
        from openpyxl.utils import get_column_letter
        wb = load_workbook(filename)
        # Ensure all main sheets are visible
        for sheet_name in ['Events', 'MDR_Texts', 'Summary']:
            if sheet_name in wb.sheetnames:
                wb[sheet_name].sheet_state = 'visible'
        # Reorder sheets using openpyxl's move_sheet
        desired_order = ['Events', 'MDR_Texts', 'Summary']
        for idx, sheet_name in enumerate(desired_order):
            if sheet_name in wb.sheetnames:
                wb.move_sheet(wb[sheet_name], offset=idx - wb.sheetnames.index(sheet_name))
        # Color sheet tabs
        tab_colors = {
            'Events': '1072BA',       # Blue
            'MDR_Texts': 'E67E22',    # Orange
            'Summary': '27AE60'       # Green
        }
        for sheet_name, color in tab_colors.items():
            if sheet_name in wb.sheetnames:
                wb[sheet_name].sheet_properties.tabColor = color
        # Premium formatting for Events sheet (the "golden sheet")
        if 'Events' in wb.sheetnames:
            ws = wb['Events']
            
            # Professional blue header (#1072BA)
            header_fill = PatternFill(start_color='1072BA', end_color='1072BA', fill_type='solid')
            header_font = Font(bold=True, name='Calibri', size=12, color='FFFFFF')
            
            # Apply premium header formatting
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
            
            # Freeze panes for better navigation
            ws.freeze_panes = 'B2'
            
            # Smart column width optimization with text wrapping
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                col_name = col[0].value
                
                # Check all cells in the column for content length (excluding header since it wraps)
                for cell in col[1:]:  # Skip header row
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            max_length = max(max_length, cell_length)
                    except:
                        pass
                
                # Set optimal width based on content type (recommended best practices)
                if col_name and any(date_word in col_name.lower() for date_word in ['date', 'received']):
                    # Date columns: 14px width (optimal for MM/DD/YYYY format)
                    ws.column_dimensions[col_letter].width = 14
                elif col_name and any(link_word in col_name.lower() for link_word in ['link', 'url']):
                    # Link columns: 30px width (shows URL structure)
                    ws.column_dimensions[col_letter].width = 30
                elif col_name and any(text_word in col_name.lower() for text_word in ['text', 'description', 'problems', 'outcome', 'treatment']):
                    # Long text columns: 25px width (for detailed content)
                    ws.column_dimensions[col_letter].width = 25
                elif col_name and any(id_word in col_name.lower() for id_word in ['id', 'key', 'number']):
                    # ID/Key columns: 15px width (perfect for identifiers)
                    ws.column_dimensions[col_letter].width = 15
                elif col_name and any(flag_word in col_name.lower() for flag_word in ['flag']):
                    # Flag columns: 16px width (good for short categorical data)
                    ws.column_dimensions[col_letter].width = 16
                else:
                    # Standard columns: 20px width (optimal for medium text)
                    ws.column_dimensions[col_letter].width = 20
            
            # Professional alternating row colors (subtle gray and white)
            fill1 = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')  # Light gray
            fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
            
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = fill1 if i % 2 == 0 else fill2
                for cell in row:
                    cell.fill = fill
            
            # Clean, professional borders
            thin = Side(border_style="thin", color="E0E0E0")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
        
        # Premium formatting for Summary sheet (green theme to match tab)
        if 'Summary' in wb.sheetnames:
            ws = wb['Summary']
            # Professional green header (#27AE60) to match tab color
            header_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
            header_font = Font(bold=True, name='Calibri', size=12, color='FFFFFF')
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
            # Freeze top row and first column
            ws.freeze_panes = 'B2'
            # Smart column width optimization
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = max(12, min(max_length + 2, 40))
            # Harmonious alternating row shading (subtle green theme)
            fill1 = PatternFill(start_color='F0F8F0', end_color='F0F8F0', fill_type='solid')  # Very light green
            fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = fill1 if i % 2 == 0 else fill2
                for cell in row:
                    cell.fill = fill
            # Add gridlines (borders) to all cells
            thin = Side(border_style="thin", color="BBBBBB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
                # After writing all summary blocks to the Summary sheet, apply cell merging for improved demographics formatting
                # Find the start row of the demographics table
                demo_header = "Patient Demographics"
                demo_start = None
                for row in ws.iter_rows():
                    if row[0].value == demo_header:
                        demo_start = row[0].row
                        break
                # Note: demo_table formatting removed as it's no longer used in the new Summary structure
        # Premium formatting for MDR_Texts sheet (orange header to match tab)
        if 'MDR_Texts' in wb.sheetnames:
            ws = wb['MDR_Texts']
            # Professional orange header (#E67E22) to match tab color
            header_fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
            header_font = Font(bold=True, name='Calibri', size=12, color='FFFFFF')
            for cell in ws[1]:
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = header_fill
            ws.freeze_panes = 'B2'
            
            # Optimized column widths for MDR_Texts
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                col_name = col[0].value
                
                # Smart width based on column content type
                if col_name and 'event_id' in col_name.lower():
                    # Event ID: 12px width (compact for IDs)
                    ws.column_dimensions[col_letter].width = 12
                elif col_name and 'text_type_code' in col_name.lower():
                    # Text Type Code: 15px width (short codes)
                    ws.column_dimensions[col_letter].width = 15
                # MDR Text Key column removed - no longer needed
                elif col_name and 'maude_report_link' in col_name.lower():
                    # MAUDE Report Link: 35px width (URLs need more space)
                    ws.column_dimensions[col_letter].width = 35
                elif col_name and 'text' in col_name.lower():
                    # Text content: 50px width (long text content)
                    ws.column_dimensions[col_letter].width = 50
                else:
                    # Default: calculate optimal width
                    max_length = 0
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    ws.column_dimensions[col_letter].width = max(12, min(max_length + 1, 30))
            fill1 = PatternFill(start_color='F7F7F7', end_color='F7F7F7', fill_type='solid')
            fill2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
                fill = fill1 if i % 2 == 0 else fill2
                for cell in row:
                    cell.fill = fill
            thin = Side(border_style="thin", color="BBBBBB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = border
        wb.save(filename)
        
        # Memory cleanup after all processing is complete
        if 'events_flat_df' in locals():
            del events_flat_df
        if 'main_fields_df' in locals():
            del main_fields_df
        import gc
        gc.collect()
        
        print(f"Export completed: {filename}")
        return filename

def export_raw_events_only():
    """Export ALL raw event data without any field exclusions - MAXIMUM PERFORMANCE"""
    import pandas as pd
    import json
    import os
    import gc
    from datetime import datetime
    
    print("Starting Raw Events export (ALL fields) - MAXIMUM PERFORMANCE...")
    
    with get_db_connection() as conn:
        # Get all events with raw JSON
        events_df = pd.read_sql_query('SELECT id, raw_json FROM events', conn)
        
        if events_df.empty:
            raise Exception("No data found in database. Please run a search first.")
        
        print(f"Processing {len(events_df)} events for Raw Events export...")
        
        # OPTIMIZATION 1: Adaptive chunking - only for large datasets
        all_events_data = []
        
        if len(events_df) > 5000:  # Only chunk for large datasets
            print("Large dataset detected - using chunked processing for memory efficiency...")
            chunk_size = 1000
            
            for chunk_start in range(0, len(events_df), chunk_size):
                chunk_end = min(chunk_start + chunk_size, len(events_df))
                chunk_df = events_df.iloc[chunk_start:chunk_end]
                
                print(f"Processing chunk {chunk_start//chunk_size + 1}/{(len(events_df)-1)//chunk_size + 1} ({chunk_start+1}-{chunk_end})")
                
                # Extract ALL fields from raw JSON (no exclusions)
                chunk_data = []
                for _, row in chunk_df.iterrows():
                    if row['raw_json']:
                        try:
                            event = json.loads(row['raw_json'])
                            # Extract ALL fields without any restrictions
                            extracted = extract_all_fields_optimized(event, row['id'])
                            chunk_data.append(extracted)
                        except Exception as e:
                            print(f"Error processing event {row['id']}: {str(e)}")
                            chunk_data.append({'event_id': row['id'], 'error': str(e)})
                
                all_events_data.extend(chunk_data)
                
                # OPTIMIZATION 2: Clear chunk data and force garbage collection
                del chunk_data
                gc.collect()
        else:
            print("Small dataset - processing all records at once for maximum speed...")
            # Process all records at once for small datasets (faster)
            for _, row in events_df.iterrows():
                if row['raw_json']:
                    try:
                        event = json.loads(row['raw_json'])
                        # Extract ALL fields without any restrictions
                        extracted = extract_all_fields_optimized(event, row['id'])
                        all_events_data.append(extracted)
                    except Exception as e:
                        print(f"Error processing event {row['id']}: {str(e)}")
                        all_events_data.append({'event_id': row['id'], 'error': str(e)})
        
        if not all_events_data:
            raise Exception("No valid data found to export.")
        
        print("Creating DataFrame...")
        # Create DataFrame with ALL fields
        events_flat_df = pd.DataFrame(all_events_data)
        print(f"Raw Events DataFrame created with {len(events_flat_df)} rows and {len(events_flat_df.columns)} columns")
        
        # Sanitize the DataFrame to prevent Excel errors
        print("Sanitizing data for Excel compatibility...")
        events_flat_df = sanitize_df(events_flat_df)
        
        # OPTIMIZATION 3: Clear original data
        del all_events_data
        gc.collect()
        
        # OPTIMIZATION 4: Skip date formatting for maximum speed
        # events_flat_df = format_all_date_columns(events_flat_df)  # COMMENTED OUT FOR SPEED
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'MAUDEMetrics_RawEvents_{timestamp}.xlsx'
        
        print("Writing to Excel with minimal formatting...")
        # OPTIMIZATION 5: Write to Excel with NO formatting for maximum speed
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            events_flat_df.to_excel(writer, sheet_name='Raw_Events', index=False)
            
            # OPTIMIZATION 6: Minimal formatting only
            wb = writer.book
            ws = wb['Raw_Events']
            
            # Only basic header formatting - no coloring
            from openpyxl.styles import Font
            header_font = Font(bold=True)
            
            for cell in ws[1]:
                cell.font = header_font
            
            # Freeze header only
            ws.freeze_panes = 'B2'
        
        # OPTIMIZATION 7: Clear DataFrame and force garbage collection
        del events_flat_df
        gc.collect()
        
        print(f"Raw Events export completed: {filename}")
        return filename

def extract_all_fields_optimized(event, event_id):
    """Extract ALL fields from raw JSON without any exclusions - OPTIMIZED VERSION"""
    result = {'event_id': event_id}
    
    def flatten_dict_optimized(d, parent_key='', sep='_'):
        """Recursively flatten nested dictionaries - OPTIMIZED"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict_optimized(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Handle arrays by creating numbered fields - OPTIMIZED
                for i, item in enumerate(v, 1):
                    if isinstance(item, dict):
                        items.extend(flatten_dict_optimized(item, f"{new_key}_{i}", sep=sep).items())
                    else:
                        # Sanitize text values to prevent Excel errors
                        if isinstance(item, str):
                            item = sanitize_text(item)
                        items.append((f"{new_key}_{i}", item))
            else:
                # Sanitize text values to prevent Excel errors
                if isinstance(v, str):
                    v = sanitize_text(v)
                items.append((new_key, v))
        return dict(items)
    
    # Flatten the entire event structure
    flattened = flatten_dict_optimized(event)
    result.update(flattened)
    
    return result

def extract_all_fields(event, event_id):
    """Extract ALL fields from raw JSON without any exclusions"""
    result = {'event_id': event_id}
    
    def flatten_dict(d, parent_key='', sep='_'):
        """Recursively flatten nested dictionaries"""
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_dict(v, new_key, sep=sep).items())
            elif isinstance(v, list):
                # Handle arrays by creating numbered fields
                for i, item in enumerate(v, 1):
                    if isinstance(item, dict):
                        items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                    else:
                        items.append((f"{new_key}_{i}", item))
            else:
                items.append((new_key, v))
        return dict(items)
    
    # Flatten the entire event structure
    flattened = flatten_dict(event)
    result.update(flattened)
    
    return result

@app.route('/', methods=['GET', 'POST'])
def index():
    with get_db_connection() as conn:
        total_events = conn.execute('SELECT COUNT(*) as count FROM events').fetchone()['count']
        is_fresh_start = (total_events == 0)
    if request.method == 'POST':
        product_code = request.form.get('product_code', '')
        brand_name = request.form.get('brand_name', '')
        device_generic_name = request.form.get('device_generic_name', '')
        start_date = request.form.get('start_date', '')
        end_date = request.form.get('end_date', '')
        max_records = request.form.get('max_records', '')
        manufacturer = request.form.get('manufacturer', '')
        # Remove manufacturer from search logic
        
        # Build the query
        base_query = "https://api.fda.gov/device/event.json?search="
        search_params = []
        if start_date and end_date:
            search_params.append(f"date_received:[{start_date}+TO+{end_date}]")
        if product_code:
            codes = [c.strip() for c in product_code.split(',') if c.strip()]
            if codes:
                code_query = ' OR '.join([f'device.device_report_product_code:{c}*' for c in codes])
                search_params.append(f'({code_query})')
        if brand_name:
            names = [n.strip() for n in brand_name.split(',') if n.strip()]
            if names:
                # Use phrase search for brand names with spaces
                name_query = ' OR '.join([f'device.brand_name:"{n}"' for n in names])
                search_params.append(f'({name_query})')
        if device_generic_name:
            generic_names = [n.strip() for n in device_generic_name.split(',') if n.strip()]
            if generic_names:
                # Use phrase search with word order variations for more precise matching
                generic_queries = []
                for name in generic_names:
                    words = name.split()
                    if len(words) > 1:
                        # Create variations for multi-word terms
                        generic_queries.append(f'device.generic_name:"{name}"')
                        # Create FDA naming convention variations (comma-separated)
                        if len(words) == 2:
                            # Two words: "word1 word2" -> "word2, word1"
                            reversed_name = ", ".join(reversed(words))
                            generic_queries.append(f'device.generic_name:"{reversed_name}"')
                        elif len(words) == 3:
                            # Three words: "word1 word2 word3" -> "word3, word1 word2" and "word2 word3, word1"
                            generic_queries.append(f'device.generic_name:"{words[2]}, {words[0]} {words[1]}"')
                            generic_queries.append(f'device.generic_name:"{words[1]} {words[2]}, {words[0]}"')
                        elif len(words) > 3:
                            # Four+ words: create common FDA variations
                            # "word1 word2 word3 word4" -> "word4, word1 word2 word3"
                            generic_queries.append(f'device.generic_name:"{words[-1]}, {" ".join(words[:-1])}"')
                            # Also try "word3 word4, word1 word2"
                            if len(words) >= 4:
                                generic_queries.append(f'device.generic_name:"{words[-2]} {words[-1]}, {" ".join(words[:-2])}"')
                    else:
                        # Single word - use phrase search
                        generic_queries.append(f'device.generic_name:"{name}"')
                generic_query = ' OR '.join(generic_queries)
                search_params.append(f'({generic_query})')
        if manufacturer:
            manufacturers = [n.strip() for n in manufacturer.split(',') if n.strip()]
            if manufacturers:
                manufacturer_query = ' OR '.join([f'device.manufacturer_d_name:{n}*' for n in manufacturers])
                search_params.append(f'({manufacturer_query})')
        if search_params:
            base_query += "+AND+".join(search_params)
        else:
            base_query += "*"  # Get all records if no filters
        
        max_records_int = None
        if max_records:
            try:
                max_records_int = int(max_records)
            except ValueError:
                pass
        
        print(f"Fetching data with query: {base_query}")
        # Fetch the first page to get the total count
        preview_query = f"{base_query}&limit=1"
        preview_response = requests.get(preview_query)
        total_count = 0
        if preview_response.status_code == 200:
            preview_data = preview_response.json()
            total_count = preview_data.get('meta', {}).get('results', {}).get('total', 0)
        session['total_count'] = total_count
        data = fetch_all_API_data(base_query, max_records_int)
        
        if data:
            print(f"Saving {len(data)} records to database...")
            save_comprehensive_data(data)
            return redirect(url_for('results'))
        else:
            return render_template('index.html', error="No results found or an error occurred.", is_fresh_start=is_fresh_start)
    return render_template('index.html', is_fresh_start=is_fresh_start)

@app.route('/results')
def results():
    with get_db_connection() as conn:
        total_events = conn.execute('SELECT COUNT(*) as count FROM events').fetchone()['count']
        total_devices = conn.execute('SELECT COUNT(*) as count FROM devices').fetchone()['count']
        total_patients = conn.execute('SELECT COUNT(*) as count FROM patients').fetchone()['count']
        is_fresh_start = (total_events == 0)
        # Get recent events
        recent_events = conn.execute('''
            SELECT e.report_number, e.event_type, e.date_received, d.manufacturer_d_name,
                   d.brand_name, d.generic_name
            FROM events e
            LEFT JOIN devices d ON e.id = d.event_id
            ORDER BY e.date_added DESC
            LIMIT 50
        ''').fetchall()
    # Format date_received as mm/dd/yyyy
    formatted_events = []
    for event in recent_events:
        event = dict(event)
        if event['date_received']:
            try:
                dt = datetime.strptime(event['date_received'], '%Y%m%d')
                event['date_received'] = dt.strftime('%m/%d/%Y')
            except Exception:
                pass
        formatted_events.append(event)
    total_count = session.get('total_count', None)
    return render_template('results.html', 
                         total_events=total_events,
                         total_devices=total_devices,
                         total_patients=total_patients,
                         recent_events=formatted_events,
                         total_count=total_count,
                         is_fresh_start=is_fresh_start)

@app.route('/export')
def export_data():
    try:
        filename = export_to_excel(include_raw_events=False)
        return send_file(filename, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return f"Error exporting data: {str(e)}", 500

@app.route('/export/raw')
def export_raw_events():
    try:
        filename = export_raw_events_only()
        return send_file(filename, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return f"Error exporting raw events: {str(e)}", 500

@app.route('/analytics')
def analytics():
    import pandas as pd
    import json
    with get_db_connection() as conn:
        total_events = conn.execute('SELECT COUNT(*) as count FROM events').fetchone()['count']
        is_fresh_start = (total_events == 0)
        # Load all raw_json from events
        events_df = pd.read_sql_query('SELECT id, raw_json FROM events', conn)
        all_events_data = []
        # Use the same field list as export_to_excel
        field_list = [
            'adverse_event_flag', 'product_problems', 'product_problem_flag', 'date_of_event', 'date_report', 'date_received',
            'device_date_of_manufacturer', 'event_type', 'number_devices_in_event', 'number_patients_in_event', 'previous_use_code',
            'remedial_action', 'removal_correction_number', 'report_number', 'single_use_flag', 'report_source_code',
            'health_professional', 'reporter_occupation_code', 'initial_report_to_fda', 'reprocessed_and_reused_flag',
            'device.device_sequence_number', 'device.device_event_key', 'device.date_received', 'device.brand_name',
            'device.generic_name', 'device.udi_di', 'device.udi_public', 'device.device_report_product_code',
            'device.model_number', 'device.catalog_number', 'device.lot_number', 'device.other_id_number',
            'device.expiration_date_of_device', 'device.device_age_text', 'device.device_availability',
            'device.date_returned_to_manufacturer', 'device.device_evaluated_by_manufacturer', 'device.device_operator',
            'device.implant_flag', 'device.date_removed_flag', 'device.manufacturer_d_name', 'device.manufacturer_d_address_1',
            'device.manufacturer_d_address_2', 'device.manufacturer_d_city', 'device.manufacturer_d_state',
                    'device.manufacturer_d_zip_code', 'device.manufacturer_d_zip_code_ext', 'device.manufacturer_d_postal_code',
        'device.manufacturer_d_country', 'device.device_class', 'device.device_name', 'device.fei_number', 'device.medical_specialty_description', 'device.registration_number', 'patient.date_received', 'patient.patient_sequence_number', 'patient.patient_age',
        'patient.patient_sex', 'patient.patient_weight', 'patient.patient_ethnicity', 'patient.patient_race',
        'patient.patient_problems', 'patient.sequence_number_outcome', 'patient.sequence_number_treatment',
        'mdr_text.date_report', 'mdr_text.mdr_text_key', 'mdr_text.patient_sequence_number', 'mdr_text.text',
        'mdr_text.text_type_code', 'type_of_report', 'date_facility_aware', 'report_date', 'report_to_fda',
        'date_report_to_fda', 'report_to_manufacturer', 'date_report_to_manufacturer', 'event_location', 'distributor_name',
        'distributor_address_1', 'distributor_address_2', 'distributor_city', 'distributor_state', 'distributor_zip_code',
        'distributor_zip_code_ext', 'manufacturer_name', 'manufacturer_address_1', 'manufacturer_address_2',
        'manufacturer_city', 'manufacturer_postal_code', 'manufacturer_state', 'manufacturer_zip_code',
        'manufacturer_zip_code_ext', 'manufacturer_country', 'manufacturer_contact_address_1',
        'manufacturer_contact_address_2', 'manufacturer_contact_area_code', 'manufacturer_contact_city',
        'manufacturer_contact_country', 'manufacturer_contact_exchange', 'manufacturer_contact_extension',
        'manufacturer_contact_t_name', 'manufacturer_contact_f_name', 'manufacturer_contact_l_name',
        'manufacturer_contact_pcity', 'manufacturer_contact_pcountry', 'manufacturer_contact_phone_number',
        'manufacturer_contact_plocal', 'manufacturer_contact_postal_code', 'manufacturer_contact_state',
        'manufacturer_contact_zip_code', 'manufacturer_contact_zip_ext', 'manufacturer_gl_name', 'manufacturer_gl_city',
        'manufacturer_gl_country', 'manufacturer_gl_postal_code', 'manufacturer_gl_state', 'manufacturer_gl_address_1',
        'manufacturer_gl_address_2', 'manufacturer_gl_zip_code', 'manufacturer_gl_zip_code_ext', 'date_manufacturer_received',
        'source_type', 'event_key', 'mdr_report_key', 'manufacturer_link_flag', 'device name', 'fei_number',
        'medical_specialty_description', 'registration_number', 'regulation_number'
        ]
        def extract_event_fields(event, field_list, event_id=None):
            result = {}
            if event_id is not None:
                result['event_id'] = event_id
            # Top-level fields
            for field in field_list:
                if '.' not in field and not field.startswith('device') and not field.startswith('patient') and not field.startswith('mdr_text'):
                    value = event.get(field, '')
                    if isinstance(value, list):
                        result[field] = ''
                        for i, v in enumerate(value):
                            result[f'{field}_{i+1}'] = v
                    else:
                        result[field] = value
            # Device fields
            devices = event.get('device', [])
            for i, device in enumerate(devices):
                for field in field_list:
                    if field.startswith('device.'):
                        subfield = field.split('.', 1)[1]
                        value = device.get(subfield, '')
                        if isinstance(value, list):
                            result[f'device_{subfield}_{i+1}'] = ''
                            for j, v in enumerate(value):
                                result[f'device_{subfield}_{i+1}_{j+1}'] = v
                        else:
                            result[f'device_{subfield}_{i+1}'] = value
            # Patient fields
            patients = event.get('patient', [])
            for i, patient in enumerate(patients):
                for field in field_list:
                    if field.startswith('patient.'):
                        subfield = field.split('.', 1)[1]
                        value = patient.get(subfield, '')
                        if isinstance(value, list):
                            result[f'patient_{subfield}_{i+1}'] = ''
                            for j, v in enumerate(value):
                                result[f'patient_{subfield}_{i+1}_{j+1}'] = v
                        else:
                            result[f'patient_{subfield}_{i+1}'] = value
            return result
        for _, row in events_df.iterrows():
            if row['raw_json']:
                try:
                    event = json.loads(row['raw_json'])
                    extracted = extract_event_fields(event, field_list, event_id=row['id'])
                    all_events_data.append(extracted)
                except Exception as e:
                    all_events_data.append({'event_id': row['id'], 'error': str(e)})
        if not all_events_data:
            # fallback: show page with no data
            return render_template('analytics.html',
                total_reports=0,
                patient_demographics=[],
                event_type_table=[],
                report_source_table=[],
                source_type_table=[],
                occupation_table=[],
                product_code_table=[],
                model_number_table=[],
                manufacturer_table=[],
                manufacturer_country_table=[],
                brand_name_table=[],
                product_problems_table=[],
                patient_problems_table=[],
                missing_patients=[])
        df = pd.DataFrame(all_events_data)
        total_reports = len(df)
        # Patient Demographics
        def extract_numeric(val):
            import re
            if pd.isnull(val):
                return None
            match = re.search(r'\d+(\.\d+)?', str(val))
            return float(match.group()) if match else None

        demo_table = []
        # Age
        age_cols = [c for c in df.columns if c.startswith('patient_patient_age')]
        ages = pd.Series([extract_numeric(v) for v in df[age_cols].values.flatten()]).dropna()
        age_val = f"{int(ages.median())} ({int(ages.min())}-{int(ages.max())})" if not ages.empty else "N/A"
        demo_table.append({"characteristic": "Age (years) median (range)", "value": age_val, "frequency": "", "percentage": ""})
        # Weight
        weight_cols = [c for c in df.columns if c.startswith('patient_patient_weight')]
        weights = pd.Series([extract_numeric(v) for v in df[weight_cols].values.flatten()]).dropna()
        weight_val = f"{weights.median():.1f} ({weights.min():.1f}-{weights.max():.1f})" if not weights.empty else "N/A"
        demo_table.append({"characteristic": "Weight median (range)", "value": weight_val, "frequency": "", "percentage": ""})
        # Sex
        sex_cols = [c for c in df.columns if c.startswith('patient_patient_sex')]
        sex_vals = pd.Series(df[sex_cols].values.flatten()).dropna()
        for i, (k, v) in enumerate(sex_vals.value_counts().items()):
            demo_table.append({
                "characteristic": "Sex" if i == 0 else "", "value": k, "frequency": v, "percentage": f"{100*v/len(sex_vals):.1f}%"})
        # Ethnicity
        eth_cols = [c for c in df.columns if c.startswith('patient_patient_ethnicity')]
        eth_vals = pd.Series(df[eth_cols].values.flatten()).dropna()
        for i, (k, v) in enumerate(eth_vals.value_counts().items()):
            demo_table.append({
                "characteristic": "Ethnicity" if i == 0 else "", "value": k, "frequency": v, "percentage": f"{100*v/len(eth_vals):.1f}%"})
        # Race
        race_cols = [c for c in df.columns if c.startswith('patient_patient_race')]
        race_vals = pd.Series(df[race_cols].values.flatten()).dropna()
        for i, (k, v) in enumerate(race_vals.value_counts().items()):
            demo_table.append({
                "characteristic": "Race" if i == 0 else "", "value": k, "frequency": v, "percentage": f"{100*v/len(race_vals):.1f}%"})
        patient_demographics = demo_table
        # Helper for event/product tables
        def make_table(df, prefix, label):
            cols = [c for c in df.columns if c.startswith(prefix)]
            vals = pd.Series(df[cols].values.flatten()).dropna()
            total = len(vals)
            counts = vals.value_counts()
            return [{"label": k, "count": v, "percent": f"{100*v/total:.1f}%"} for k, v in counts.items()] if total > 0 else []
        # Event/Product Characteristics
        event_type_table = make_table(df, 'event_type', 'Event Type')
        report_source_table = make_table(df, 'report_source_code', 'Report Source')
        source_type_table = make_table(df, 'source_type', 'Source Type')
        occupation_table = make_table(df, 'reporter_occupation_code', 'Reporter Occupation Code')
        product_code_table = make_table(df, 'device_device_report_product_code', 'Product Code')
        model_number_table = make_table(df, 'device_model_number', 'Model Number')
        manufacturer_table = make_table(df, 'device_manufacturer_d_name', 'Manufacturer')
        manufacturer_country_table = make_table(df, 'device_manufacturer_d_country', 'Manufacturer Country')
        brand_name_table = make_table(df, 'device_brand_name', 'Brand Name')
        # Type of Device (Generic Name)
        generic_name_table = make_table(df, 'device_generic_name', 'Product Class')
        # Device Problem
        prod_cols = [c for c in df.columns if c.startswith('product_problems')]
        prod_vals = pd.Series(df[prod_cols].values.flatten()).dropna()
        total_prod = len(prod_vals)
        prod_counts = prod_vals.value_counts()
        product_problems_table = [{"label": k, "count": v, "percent": f"{100*v/total_prod:.1f}%"} for k, v in prod_counts.items()] if total_prod > 0 else []
        # Patient Problem
        pprob_cols = [c for c in df.columns if c.startswith('patient_patient_problems')]
        pprob_vals = pd.Series(df[pprob_cols].values.flatten()).dropna()
        total_pprob = len(pprob_vals)
        pprob_counts = pprob_vals.value_counts()
        patient_problems_table = [{"label": k, "count": v, "percent": f"{100*v/total_pprob:.1f}%"} for k, v in pprob_counts.items()] if total_pprob > 0 else []
        # Events missing patient data (as before)
        missing_patients = conn.execute('''
            SELECT e.id as event_id, e.report_number
            FROM events e
            LEFT JOIN patients p ON e.id = p.event_id
            WHERE p.id IS NULL
        ''').fetchall()
        # --- Dynamic Chart Data Preparation ---
        # 1. Event Types per Brand Name (top 10 brands)
        brand_cols = [c for c in df.columns if c.startswith('device_brand_name')]
        eventtype_cols = [c for c in df.columns if c.startswith('event_type')]
        brand_eventtype_pairs = []
        for i in range(len(df)):
            brands = [df.iloc[i][col] for col in brand_cols if pd.notnull(df.iloc[i][col]) and str(df.iloc[i][col]).strip()]
            event_types = [df.iloc[i][col] for col in eventtype_cols if pd.notnull(df.iloc[i][col]) and str(df.iloc[i][col]).strip()]
            for b in brands:
                for e in event_types:
                    brand_eventtype_pairs.append((b, e))
        import collections
        brand_counts = collections.Counter([b for b, _ in brand_eventtype_pairs])
        top_brands = [b for b, _ in brand_counts.most_common(10)]
        filtered_pairs = [(b, e) for b, e in brand_eventtype_pairs if b in top_brands]
        # Build a nested dict: {brand: {event_type: count}}
        brand_eventtype_dict = collections.defaultdict(lambda: collections.Counter())
        for b, e in filtered_pairs:
            brand_eventtype_dict[b][e] += 1
        # Convert to list of dicts for JSON
        brand_eventtype_data = []
        for b in top_brands:
            entry = {'brand': b}
            entry.update(brand_eventtype_dict[b])
            brand_eventtype_data.append(entry)
        # Pass as JSON
        import json as _json
    return render_template('analytics.html',
        total_reports=total_reports,
        patient_demographics=patient_demographics,
        event_type_table=event_type_table,
        report_source_table=report_source_table,
        source_type_table=source_type_table,
        occupation_table=occupation_table,
        product_code_table=product_code_table,
        model_number_table=model_number_table,
        manufacturer_table=manufacturer_table,
        manufacturer_country_table=manufacturer_country_table,
        brand_name_table=brand_name_table,
        generic_name_table=generic_name_table,
        product_problems_table=product_problems_table,
        patient_problems_table=patient_problems_table,
        missing_patients=missing_patients,
        is_fresh_start=is_fresh_start,
        brand_eventtype_data=_json.dumps(brand_eventtype_data)
    )

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/clear_data', methods=['POST'])
def clear_data():
    with get_db_connection() as conn:
        conn.execute('DELETE FROM mdr_texts')
        conn.execute('DELETE FROM patients')
        conn.execute('DELETE FROM devices')
        conn.execute('DELETE FROM events')
        conn.commit()
    session['total_count'] = None  # Reset the total_count for the results page
    return redirect(url_for('index'))



if __name__ == "__main__":
    init_db()
    app.run(host='0.0.0.0', port=5005, debug=True)
